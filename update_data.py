#!/usr/bin/env python3
"""
毎月の粗利分析Excelファイルを読み込み、index.htmlのデータを更新するスクリプト

使い方:
    python update_data.py 粗利分析_202504_1.xlsx
"""

import sys
import json
import re
from pathlib import Path
import openpyxl

def parse_excel(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    
    # シート確認
    required = ['粗利分析サマリー', '部門別明細']
    for s in required:
        if s not in wb.sheetnames:
            raise ValueError(f"シート '{s}' が見つかりません")
    
    summary_ws = wb['粗利分析サマリー']
    detail_ws  = wb['部門別明細']
    
    # =============================================
    # サマリーシート解析
    # =============================================
    DEPT_CAT = {
        '白銅': '3PL', 'ムトウユニパック': '3PL', '化研マテリアル': '3PL',
        '吉川紙商事': '3PL', 'ナチュラジャパン': '3PL',
        'trackerr': 'SaaS', '配録': 'SaaS', 'イツクルLOGI': 'SaaS',
        'KURAud': 'SaaS', 'HaKoPo': 'SaaS',
        '受託開発': '受託開発',
        'コンサルティング（青和向け）': 'コンサルティング',
    }
    SKIP = {'【3PL】','【ソフトウェア】','SaaS','受託開発','【コンサルティング】',
            '3PL 小計','SaaS 小計','ソフトウェア 小計','全部門 合計','カテゴリ','項目'}
    
    summary = []
    commission = {}
    
    rows = list(summary_ws.iter_rows(values_only=True))
    
    # タイトル行から期間取得
    title = str(rows[0][0] or '')
    year_match = re.search(r'(\d{4})年度', title)
    year = year_match.group(1) if year_match else '2025'
    
    for i, r in enumerate(rows):
        if not r or not r[0]:
            continue
        dept = str(r[0]).strip()
        
        # コミッション行
        if dept == '対象粗利（今期）':
            commission['q1_gross_current'] = int(r[1] or 0)
            commission['q2_gross_current'] = int(r[2] or 0)
        elif dept == '対象粗利（前年）':
            commission['q1_gross_prev'] = int(r[1] or 0)
            commission['q2_gross_prev'] = int(r[2] or 0)
        elif dept == '大家さんコミッション':
            pass  # 自動計算
        
        if dept in SKIP or dept.startswith('【') or dept == 'カテゴリ':
            continue
        cat = DEPT_CAT.get(dept)
        if not cat:
            continue
        
        def safe_int(v):
            if v is None or v == '-' or v == '': return 0
            try: return int(float(str(v).replace(',', '')))
            except: return 0
        
        summary.append({
            'dept': dept,
            'category': cat,
            'q1_2025': safe_int(r[1]),
            'q1_2024': safe_int(r[2]),
            'q2_2025': safe_int(r[5]),
            'q2_2024': safe_int(r[6]),
        })
    
    # =============================================
    # 明細シート解析
    # =============================================
    detail = []
    detail_rows = list(detail_ws.iter_rows(values_only=True))
    
    for r in detail_rows[3:]:  # ヘッダー3行スキップ
        if not r or not r[0]:
            continue
        period = str(r[2] or '').strip()
        if not period.startswith('2025'):
            continue
        
        def safe_float(v):
            if v is None or v == '-' or v == '': return 0.0
            try: return float(str(v).replace(',',''))
            except: return 0.0
        
        detail.append({
            'category': str(r[0]).strip(),
            'dept': str(r[1]).strip(),
            'period': period,
            'revenue': int(safe_float(r[3])),
            'cost': int(safe_float(r[4])),
            'gross': int(safe_float(r[5])),
            'margin': round(safe_float(r[6]), 6),
            'prev': int(safe_float(r[7])) if r[7] != '-' else 0,
        })
    
    # コミッション補完
    if not commission:
        commission = {
            'q1_gross_current': sum(s['q1_2025'] for s in summary),
            'q2_gross_current': sum(s['q2_2025'] for s in summary),
            'q1_gross_prev':    sum(s['q1_2024'] for s in summary),
            'q2_gross_prev':    sum(s['q2_2024'] for s in summary),
        }
    commission['ooya_rate']   = 0.13
    commission['sasaki_rate'] = 0.01
    
    from datetime import date
    return {
        'period': f'{year} Q1・Q2',
        'updatedAt': date.today().strftime('%Y年%m月%d日'),
        'summary': summary,
        'detail': detail,
        'commission': commission,
    }


def update_html(data, html_path='index.html'):
    html_path = Path(html_path)
    if not html_path.exists():
        raise FileNotFoundError(f'{html_path} が見つかりません')
    
    content = html_path.read_text(encoding='utf-8')
    
    # DEFAULT_DATA ブロックを置換
    new_data_js = f"const DEFAULT_DATA = {json.dumps(data, ensure_ascii=False, indent=2)};"
    
    pattern = r'const DEFAULT_DATA = \{[\s\S]*?\};'
    updated = re.sub(pattern, new_data_js, content, count=1)
    
    if updated == content:
        raise ValueError('DEFAULT_DATA の置換に失敗しました')
    
    html_path.write_text(updated, encoding='utf-8')
    print(f'✅ {html_path} を更新しました')
    print(f'   期間: {data["period"]}')
    print(f'   更新日: {data["updatedAt"]}')
    print(f'   部門数: {len(data["summary"])}')


def main():
    if len(sys.argv) < 2:
        print('使い方: python update_data.py <xlsxファイルパス>')
        sys.exit(1)
    
    xlsx_path = sys.argv[1]
    html_path = sys.argv[2] if len(sys.argv) > 2 else 'index.html'
    
    print(f'📂 読み込み中: {xlsx_path}')
    data = parse_excel(xlsx_path)
    
    print(f'📝 HTMLを更新中: {html_path}')
    update_html(data, html_path)
    
    print()
    print('次のステップ:')
    print('  git add index.html')
    print(f'  git commit -m "データ更新: {data[\"period\"]}"')
    print('  git push origin main')


if __name__ == '__main__':
    main()
